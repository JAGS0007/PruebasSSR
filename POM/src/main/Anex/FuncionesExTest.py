import openpyxl

class Funexcel():
    def __init__(self, driver):
        self.driver = driver

    def getRowCount(file,path,sheetName):
        Worbook = openpyxl.load_workbook(path)
        sheet = Worbook[sheetName]
        return (sheet.max_row)

    def getColumnCount(file,sheetName):
        Worbook = openpyxl.load_workbook(file)
        sheet = Worbook[sheetName]
        return (sheet.max_row)

    def readData(file,path,sheetName,rownum,columnno):
        Worbook = openpyxl.load_workbook(path)
        sheet =  Worbook[sheetName]
        return sheet.cell(row=rownum, column=columnno).value

    def writeData(file,path,sheetName,rownum,columnno,data):
        Worbook = openpyxl.load_workbook(path)
        sheet = Worbook[sheetName]
        sheet.cell(row=rownum, column=columnno).value = data
        Worbook.save(path)